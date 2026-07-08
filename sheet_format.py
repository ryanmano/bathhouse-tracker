"""Friendly spreadsheet formatting: database rows -> simple, readable rows.

All times are shown in US-Eastern (the venues' timezone). Class times render
as ranges like "9-10am" / "9:30-10:45am"; end times are derived from each
platform's raw payload (Trybe stores an end timestamp; Mariana Tek and Arketa
store a duration in minutes).
"""
from __future__ import annotations

import re
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

EASTERN = ZoneInfo("America/New_York")

# Columns actually written to the simple sheets. simple_row() computes a few
# extras (notes) that stay available if ever re-added.
SIMPLE_COLUMNS = [
    "brand", "location", "class", "date", "time",
    "observed", "price", "spots_left", "capacity",
]

# Pre-start summary: one row per class = its final reading before it began.
# The daily file is split into one tab per brand, so the brand column is
# dropped from the per-tab layout (the tab name says the brand).
PRESTART_COLUMNS = [
    "brand", "location", "class", "date", "time",
    "price", "spots_left", "spots_booked", "capacity", "read_at", "mins_before",
]
# Each tab is one brand+location, so brand and location are dropped per tab.
PRESTART_SHEET_COLUMNS = [
    "class", "date", "time",
    "price", "spots_left", "spots_booked", "capacity", "read_at", "mins_before",
]

# Brand display order for tab arrangement.
BRAND_ORDER = ["bathhouse", "othership", "lore"]

# Running overview: one row per club per day with that day's final totals.
SUMMARY_COLUMNS = [
    "club", "date", "spots_left", "spots_booked", "capacity", "price",
]

HEADERS = {
    "brand": "Brand", "location": "Location", "club": "Club", "class": "Class",
    "date": "Date", "time": "Time", "price": "Price", "spots_left": "Spots Left",
    "spots_booked": "Spots Booked", "capacity": "Total Spots",
    "notes": "Notes", "observed": "Observed",
    "read_at": "Last Read", "mins_before": "Min Before Start",
}

COLUMN_WIDTHS = {
    "brand": 13, "location": 16, "club": 26, "class": 48, "date": 13, "time": 15,
    "price": 12, "spots_left": 13, "spots_booked": 16, "capacity": 15,
    "observed": 18, "read_at": 18, "mins_before": 18,
}

# Short columns look best centered; the wide text columns stay left-aligned.
CENTERED_COLUMNS = {
    "date", "time", "price", "spots_left", "spots_booked", "capacity",
    "observed", "read_at", "mins_before",
}

# Extra PostgREST select expressions that pull end-time ingredients out of the
# stored raw json (missing fields simply come back null).
EXTRA_SELECTS = [
    "raw_end:raw->>end_time",                 # trybe: ISO end timestamp
    "raw_dur:raw->>duration",                 # arketa: minutes
    "raw_ct_dur:raw->class_type->>duration",  # mariana_tek: minutes
]


def _parse(ts) -> datetime | None:
    if not ts:
        return None
    s = str(ts).replace("Z", "+00:00")
    # Normalize odd fractional-second widths (e.g. "...4935+00:00") so this
    # also works on Python 3.9's stricter fromisoformat, not just CI's 3.12.
    m = re.match(
        r"^(\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2})(?:\.(\d+))?([+-]\d{2}:\d{2})?$", s
    )
    if m:
        base, frac, off = m.group(1), m.group(2), m.group(3) or ""
        s = base + (f".{frac.ljust(6, '0')[:6]}" if frac else "") + off
    try:
        return datetime.fromisoformat(s)
    except ValueError:
        return None


def _minutes(val) -> int | None:
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


def _clock(dt: datetime, with_ampm: bool = True) -> str:
    hour = dt.strftime("%I").lstrip("0")
    out = hour if dt.minute == 0 else f"{hour}:{dt.minute:02d}"
    return out + (dt.strftime("%p").lower() if with_ampm else "")


def _time_range(start: datetime, end: datetime | None) -> str:
    if end is None or end <= start:
        return _clock(start)
    same_half = start.strftime("%p") == end.strftime("%p")
    return f"{_clock(start, not same_half)}-{_clock(end)}"


def _money(price) -> str:
    if price is None or price == "":
        return ""
    p = float(price)
    return f"${int(p)}" if p == int(p) else f"${p:.2f}"


def _clean_class(name, brand) -> str:
    """Trim class names down to the useful part.

    The Brand column already says the venue, so redundant prefixes, leading
    emoji, and trailing "- NN min" durations are stripped. Bathhouse only
    sells the day pass, so its verbose booking label collapses to "Day Pass".
    """
    if not name:
        return ""
    s = re.sub(r"^[^\w$]+", "", str(name)).strip()  # leading emoji/symbols
    if "Day Pass" in s:
        return "Day Pass"
    s = re.sub(r"^Lore Bathing Club\s*[-–]\s*", "", s)  # redundant brand prefix
    s = re.sub(r"\s*[-–]\s*\d+\s*min\.?$", "", s, flags=re.I)  # trailing duration
    return s.strip() or str(name).strip()


def simple_row(row: dict) -> dict:
    start = _parse(row.get("start_time"))
    end = _parse(row.get("raw_end"))
    if end is None and start is not None:
        mins = _minutes(row.get("raw_ct_dur")) or _minutes(row.get("raw_dur"))
        if mins:
            end = start + timedelta(minutes=mins)
    start_et = start.astimezone(EASTERN) if start else None
    end_et = end.astimezone(EASTERN) if end else None
    observed_et = None
    observed = _parse(row.get("observed_at"))
    if observed:
        observed_et = observed.astimezone(EASTERN)

    notes = row.get("price_tier") or ""
    if row.get("is_waitlist") in (True, 1, "true", "True"):
        notes = f"{notes}; FULL" if notes else "FULL"

    return {
        "brand": (row.get("brand") or "").capitalize(),
        "location": row.get("location") or "",
        "class": _clean_class(row.get("class_name"), row.get("brand")),
        "date": f"{start_et.strftime('%a %b')} {start_et.day}" if start_et else "",
        "time": _time_range(start_et, end_et) if start_et else "",
        "price": _money(row.get("price")),
        "spots_left": row.get("spots_available"),
        "spots_booked": row.get("spots_booked"),
        "capacity": row.get("capacity"),
        "notes": notes,
        "observed": f"{observed_et.strftime('%b')} {observed_et.day}, {_clock(observed_et)}"
        if observed_et
        else "",
    }


def simple_rows(rows: list[dict]) -> list[dict]:
    """Sort by class start (then brand/location/observation) and format."""
    ordered = sorted(
        rows,
        key=lambda r: (
            r.get("start_time") or "",
            r.get("brand") or "",
            str(r.get("location") or ""),
            r.get("observed_at") or "",
        ),
    )
    return [simple_row(r) for r in ordered]


def prestart_rows(rows: list[dict], now: datetime | None = None) -> list[dict]:
    """One row per class: its final reading taken at/before the class start.

    Only classes that have already started are included (their pre-start reading
    is final). `mins_before` is how many minutes before start that reading was
    taken — small numbers mean we caught the last-minute booking state.
    """
    now = now or datetime.now(timezone.utc)
    best: dict[tuple, tuple[datetime, datetime, dict]] = {}
    for r in rows:
        start = _parse(r.get("start_time"))
        obs = _parse(r.get("observed_at"))
        if start is None or obs is None or start >= now or obs > start:
            continue
        key = (r.get("brand"), r.get("session_id"))
        if key not in best or obs > best[key][0]:
            best[key] = (obs, start, r)

    out: list[tuple[datetime, dict]] = []
    for obs, start, r in best.values():
        rec = simple_row(r)
        rec["read_at"] = rec.pop("observed")
        rec["mins_before"] = round((start - obs).total_seconds() / 60)
        out.append((start, rec))
    out.sort(key=lambda pair: pair[0])
    return [rec for _, rec in out]


def _totals_row(recs: list[dict]) -> dict:
    """Day totals for one location: summed spots + weighted-average price.

    The average is revenue-weighted — sum(price * spots_booked) / sum(spots_booked)
    — i.e. the average dollars actually spent per booked ticket, not the plain
    average of listed prices. (Uses each session's final pre-start price as a
    proxy for what its tickets sold at.) Blank when no priced bookings exist,
    as with Othership, whose price isn't public.
    """
    row = {c: "" for c in PRESTART_SHEET_COLUMNS}
    row["class"] = "TOTALS"
    left = booked = cap = 0
    has_left = has_booked = has_cap = False
    revenue = 0.0
    tickets_priced = 0
    for r in recs:
        sl, sb, c = r.get("spots_left"), r.get("spots_booked"), r.get("capacity")
        if isinstance(sl, (int, float)):
            left += sl; has_left = True
        if isinstance(sb, (int, float)):
            booked += sb; has_booked = True
        if isinstance(c, (int, float)):
            cap += c; has_cap = True
        price = r.get("price")
        if isinstance(price, str) and price.strip() and isinstance(sb, (int, float)) and sb > 0:
            try:
                pv = float(price.replace("$", "").replace(",", ""))
            except ValueError:
                pv = 0.0
            # Exclude $0 (member-included / free) tickets from the average —
            # they aren't a dollars-spent signal.
            if pv > 0:
                revenue += pv * sb
                tickets_priced += sb
    if has_left:
        row["spots_left"] = left
    if has_booked:
        row["spots_booked"] = booked
    if has_cap:
        row["capacity"] = cap
    if tickets_priced > 0:
        row["price"] = _money(revenue / tickets_priced)
    return row


def _style_sheet(ws, columns: list[str], rows: list[dict], totals: dict | None = None) -> None:
    """Fill one worksheet: bold banded header, frozen top row, filters, widths.

    When `totals` is given it is appended as a bold, shaded summary row that is
    kept outside the filter range.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    left_align = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    ws.append([HEADERS[c] for c in columns])
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for rec in rows:
        ws.append([rec.get(c) for c in columns])
    last_data_row = ws.max_row  # header + data, before any totals row

    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = COLUMN_WIDTHS.get(col, 12)
        align = center if col in CENTERED_COLUMNS else left_align
        for cell in ws[letter][1:]:  # data rows only; header keeps its styling
            cell.alignment = align
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{last_data_row}"

    if totals is not None:
        ws.append([totals.get(c) for c in columns])
        tr = ws.max_row
        totals_fill = PatternFill("solid", fgColor="E8EEF7")
        top_border = Border(top=Side(style="thin", color="1F3A5F"))
        for i, col in enumerate(columns, start=1):
            cell = ws.cell(tr, i)
            cell.font = Font(bold=True)
            cell.fill = totals_fill
            cell.border = top_border
            cell.alignment = center if col in CENTERED_COLUMNS else left_align


def _write_xlsx(rows: list[dict], columns: list[str], title: str) -> bytes:
    """Single-sheet styled workbook."""
    import io

    from openpyxl import Workbook

    wb = Workbook()
    _style_sheet(wb.active, columns, rows)
    wb.active.title = title
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_bytes(rows: list[dict]) -> bytes:
    """Full time-series sheet: every reading, one row each."""
    return _write_xlsx(simple_rows(rows), SIMPLE_COLUMNS, "Sessions")


def _tab_name(brand: str, location: str) -> str:
    """Excel-safe tab title, e.g. 'Bathhouse Atlantic Ave' (max 31 chars)."""
    name = f"{(brand or '').capitalize()} {location or ''}".strip()
    for ch in r"[]:*?/\\":
        name = name.replace(ch, " ")
    return name[:31] or "Sheet"


def prestart_xlsx_bytes(rows: list[dict]) -> bytes:
    """Clean summary workbook: one tab per brand+location, with a TOTALS row."""
    import io

    from openpyxl import Workbook

    recs = prestart_rows(rows)
    groups: dict[tuple[str, str], list[dict]] = {}
    for r in recs:
        groups.setdefault((r.get("brand") or "", r.get("location") or ""), []).append(r)

    order = {b: i for i, b in enumerate(BRAND_ORDER)}
    keys = sorted(groups, key=lambda k: (order.get(k[0].lower(), 99), k[1]))

    wb = Workbook()
    first = True
    for brand, location in keys:
        group = groups[(brand, location)]
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = _tab_name(brand, location)
        _style_sheet(ws, PRESTART_SHEET_COLUMNS, group, totals=_totals_row(group))

    if first:  # no started classes at all — keep a valid, empty workbook
        wb.active.title = "Pre-start"
        _style_sheet(wb.active, PRESTART_SHEET_COLUMNS, [])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def summary_rows(rows: list[dict], earliest: str | None = None) -> list[dict]:
    """One row per club per day: that day's final totals (spots + weighted price).

    Grouped by brand+location and Eastern calendar day, ordered by club then
    date. Days before `earliest` (YYYY-MM-DD) are skipped.
    """
    groups: dict[tuple[str, str, str], list[dict]] = {}
    for r in rows:
        st = _parse(r.get("start_time"))
        if st is None:
            continue
        ymd = st.astimezone(EASTERN).date().isoformat()
        if earliest and ymd < earliest:
            continue
        groups.setdefault((r.get("brand") or "", r.get("location") or "", ymd), []).append(r)

    order = {b: i for i, b in enumerate(BRAND_ORDER)}
    keyed = sorted(groups, key=lambda k: (order.get(k[0].lower(), 99), k[1], k[2]))

    out: list[dict] = []
    for brand, location, ymd in keyed:
        recs = prestart_rows(groups[(brand, location, ymd)])
        if not recs:
            continue
        totals = _totals_row(recs)
        d = datetime.fromisoformat(ymd)
        out.append({
            "club": _tab_name(brand, location),
            "date": f"{d.strftime('%a %b')} {d.day}",
            "spots_left": totals["spots_left"],
            "spots_booked": totals["spots_booked"],
            "capacity": totals["capacity"],
            "price": totals["price"],
        })
    return out


def summary_xlsx_bytes(rows: list[dict], earliest: str | None = None) -> bytes:
    """Single 'Summary' sheet: club × day final totals, pre-start styling."""
    return _write_xlsx(summary_rows(rows, earliest), SUMMARY_COLUMNS, "Summary")
